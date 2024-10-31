import os
import sys
import glob
import sqlite3
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt

plt.style.use('seaborn-v0_8-whitegrid')
# print('- '+'\n- '.join(plt.style.available))

# #
# I/O Functions
# #

def execute_read_script(sql_file: str, db_file: str = None) -> pd.DataFrame:
    """Run a SQL file in sqlite3."""
    # Default database file to: `./chinook_db/chinook.db``
    db_file = db_file or './chinook_db/chinook.db'

    # Connect to sqlite database
    with sqlite3.connect(db_file) as conn:
        with open(sql_file, 'r') as f:
            df = pd.read_sql(f.read(), con=conn)
    return df
        
def write_to_csv(df: pd.DataFrame, csv_file: str, verbose: bool = False):
    """Writes a pandas.DataFrame to a csv file."""
    try:
        if verbose:
            ui = input(f'Press Enter to write to csv file: {csv_file} ...')
            if ui:
                df.to_csv(csv_file,)
                print('Success!')
        else:
            # print(f'Writing to SQL script {} to CSV file {} ... ')
            df.to_csv(csv_file,)
            print('Success!')

    except Exception as err:
        print('Failed!')
        print(err)
        return

def write_to_spreadsheet(df: pd.DataFrame, sheet_name: str, xlsx_file: str) -> str:
    """Writes a pandas dataframe to an excel worksheet.
    
    Args:
        df (pd.DataFrame): Dataframe to write to excel.
        sheet_name (str): Name of the sheet to write to.
        xlsx_file (str): Path to the excel file.
    """
    # Try to load the workbook. If it doesn't exist, create a new one.
    try:
        wb = openpyxl.load_workbook(xlsx_file)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # If the sheet already exists, remove it
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    # Create the sheet
    wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    
    # Write headers to first row
    ws.append([*df.columns])
    
    # Iterate over the rows in the dataframe
    for idx, row in df.iterrows():
        # Write each row to sheet
        ws.append(row.tolist())

    # Save the workbook
    wb.save(xlsx_file)

    return xlsx_file

# # 
# Plotting Functions
# #

def create_chart1(df: pd.DataFrame, chart_file: str,):
    """Create chart for `query_1.sql` script."""
    ## Load dataframe
    # df = pd.read_csv(csv_file)
    
    # Create pivot table
    df_pivot = pd.pivot(df, index='Date', columns='Employee', values='Sales')
    df_pivot.fillna(0, inplace=True)
    df_pivot['Grand Total'] = df_pivot.apply(sum, axis=1)
    print(df_pivot)
    # print()
    # print(df_pivot.apply(sum, axis=1))
    input()

    # plot pivot table
    fig, ax = plt.subplots()

    x = df_pivot.index
    y = df_pivot[df_pivot.columns[-1]]
    ax.bar(x=x, height=y, alpha=0.25, label='Grand Total')
    
    x = df_pivot.index
    ys = df_pivot[df_pivot.columns[:-1]]
    ax.plot(x, ys, '.-')
    ax.set_xlabel('Date')
    ax.set_ylabel('Sales (in USD)')
    ax.set_
    ax.set_title('Employee Monthly Sales')
    ax.legend(loc='best')

    plt.show()

    ui = input(f'Press Enter to save chart to: {chart_file} ...')
    if not not ui:
        plt.savefig(chart_file)

    return chart_file

def create_chart2(df: pd.DataFrame, chart_file: str,):
    """Create chart for `query_2.sql` script."""
    pass

def create_chart3(df: pd.DataFrame, chart_file: str,):
    """Create chart for `query_3.sql` script."""
    pass

def create_chart4(df: pd.DataFrame, chart_file: str,):
    """Create chart for `query_4.sql` script."""
    pass

def create_charts(df: pd.DataFrame, chart_file: str,):
    """Create"""
    print(f'Creating chart: {chart_file}')
    if 'query_1' in chart_file:
        create_chart1(df, chart_file)
    elif 'query_2' in chart_file:
        create_chart2(df, chart_file)
    elif 'query_3' in chart_file:
        create_chart3(df, chart_file)
    elif 'query_4' in chart_file:
        create_chart4(df, chart_file)
    else:
        print('Error!')
    print()

# #
# Application Functions
# #

def run_project_queries(data_dir: str = None, *, verbose: bool = False):
    """Run the project queries."""
    # Default data directory to: `./data`
    data_dir = data_dir or './data'

    # Iterate over sql files to read
    for sql_file in sorted(glob.glob(f'{data_dir}/sql/*.sql')):
        # Fetch filepaths first
        csv_file = sql_file.replace('sql', 'csv')
        sheet_name = os.path.basename(sql_file).removesuffix('.sql')
        xlsx_file = os.path.dirname(sql_file).replace('sql','xlsx')+'/sql_project.xlsx'
        chart_file = f'charts/{sheet_name}.png'
        
        ## Load and save data
        # If CSV file exists, load from csv
        if os.path.exists(csv_file):
            # Load CSV data
            print(f'Loading from csv file: {csv_file}')
            df = pd.read_csv(csv_file)
            print()
        
        # If not, load from sql
        else:
            # Load SQL data into dataframe
            print(f'Loading from SQL script: {sql_file}')
            df = execute_read_script(sql_file)
            print()
            
            # Write dataframe to csv file
            print(f'Writing to csv file: {csv_file}')
            write_to_csv(df, csv_file)
            print()

        ## Create new sheet in spreadsheet
        # If excel workbook exists, skip
        if os.path.exists(xlsx_file):
            print(f'Spreadsheet `{xlsx_file}` already exists.\nSkipping new sheet: {sheet_name} ...')
            print()
            
        else:
            # Write dataframe to xlsx sheet
            print(f'Writing to spreadsheet: {xlsx_file}')
            print(f'Sheet name: {sheet_name}')
            write_to_spreadsheet(df, sheet_name, xlsx_file)
            print()

    
        # # If chart file exists, overwrite file
        # create_charts(df, chart_file)


def main():
    """Run the application."""
    run_project_queries()
    print('Done!')



if __name__ == '__main__':
    print()
    main()
    print()